"""
Generate revision_summary_short.xlsx — one row per reviewer point,
with: ID, Phase, Reviewer, Question summary, Our answer, Status, Confidence.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUT_PATH = "/Users/jysong/Desktop/Gore_lab/Sequencing/Coalescence_session_20230404/Figure_generate/Draft/v4/latex/revision/revision_response_summary.xlsx"

# ── colour palette ──────────────────────────────────────────────────────────
PHASE_COLORS = {
    "Phase 1": "D6EAF8",  # light blue
    "Phase 2": "D5F5E3",  # light green
    "Phase 3": "FEF9E7",  # light yellow
    "Phase 4": "FDEDEC",  # light red/pink
    "Phase 5": "F5EEF8",  # light purple
    "Phase 6": "FDFEFE",  # near-white
}
STATUS_COLORS = {
    "COMPLETE":     "A9DFBF",   # green
    "IN PROGRESS":  "FAD7A0",   # orange
    "NOT STARTED":  "F5CBA7",   # peach
    "DONE":         "A9DFBF",
    "PENDING DECISION": "D7BDE2",
}
HEADER_FILL  = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(size=10)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── data ────────────────────────────────────────────────────────────────────
# columns: id, phase, reviewer, question_summary, our_answer, status, confidence
ROWS = [
    # ── PHASE 1 ─────────────────────────────────────────────────────────────
    ("P1.1", "Phase 1", "R2",
     "Typo 'generalis ability' in manuscript.",
     "Already corrected in v4 — no further change needed.",
     "COMPLETE", "100%"),

    ("P1.2", "Phase 1", "R1",
     "SI p.9 references wrong Extended Data figure numbers.",
     "Fixed two files: simulations.tex (ED Fig 5→4) and pairwise_selection_correlation.tex (ED Fig 6→5, 7→6).",
     "COMPLETE", "100%"),

    ("P1.3", "Phase 1", "R2",
     "Inconsistent terminology and notation throughout manuscript.",
     "Reviewed all .tex files — usage of 'coalescence event/experiment', μ, and α_ij is already consistent; no changes required.",
     "COMPLETE", "95%"),

    ("P1.4", "Phase 1", "R3",
     "Term 'interaction strength' is ambiguous between model (μ) and experimental context.",
     "Added disambiguation sentence in results.tex at first use of μ, distinguishing model coefficient from experimental proxy (failed invasions).",
     "COMPLETE", "90%"),

    ("P1.5", "Phase 1", "R1",
     "Claim that Dominance is 'robust across similarity metrics' (line 134) overstates the result.",
     "Softened to 'broadly consistent across several metrics, but relative ranking sensitive to metric choice (Jensen–Shannon, Jaccard)'.",
     "COMPLETE", "95%"),

    ("P1.6", "Phase 1", "R2",
     "gLV model should be explicitly framed as phenomenological, not mechanistic.",
     "Added explicit sentence: 'We emphasize that the gLV model serves as a phenomenological framework… rather than as a mechanistic model.'",
     "COMPLETE", "95%"),

    ("P1.7", "Phase 1", "R1",
     "Emphasise 'cohesion without cooperation' concept, referencing Tikhonov.",
     "Expanded Tikhonov paragraph in discussion.tex to 3 sentences explaining how purely competitive interactions yield community-level cohesion.",
     "COMPLETE", "90%"),

    ("P1.8", "Phase 1", "R3",
     "Claims about natural communities (lines 292–294) are too strong; facilitation not modelled.",
     "Softened language, acknowledged 'greater taxonomic diversity, more complex interactions, or facilitative interactions not captured' by our model.",
     "COMPLETE", "90%"),

    ("P1.9", "Phase 1", "R2",
     "Discuss alternative mechanisms for community-level selection (environmental filtering, hitchhiking, shared traits).",
     "New paragraph in discussion.tex addresses all three alternatives and argues against each using pairwise correlation controls and gLV reproduction.",
     "COMPLETE", "85%"),

    ("P1.10", "Phase 1", "R2",
     "Acknowledge natural community pre-selection during stabilization phase.",
     "New paragraph in discussion.tex: 'stabilisation may pre-select species toward culture conditions, potentially reducing effective diversity' — acknowledged as limitation.",
     "COMPLETE", "85%"),

    ("P1.11", "Phase 1", "R2",
     "Nutrient enrichment ≠ increase in interaction strength; reframe this mapping.",
     "Revised Results and Discussion; added Duan 2025 citation; changed 'mean interaction strengths' to 'mean competition coefficients'; added note that the mapping is 'necessarily approximate'.",
     "COMPLETE", "80%"),

    ("P1.12", "Phase 1", "R2",
     "Specify pH measurement method in Methods.",
     "Already specified in methods.tex (Apera Instruments PH5500); no change needed.",
     "COMPLETE", "95%"),

    ("P1.13", "Phase 1", "R2",
     "Clarify biological meaning of gLV interaction coefficients (α_ij).",
     "Added biological interpretation in supplementary_methods.tex: α_ij encodes net per-capita effect including resource competition, interference, metabolite-mediated inhibition; facilitation not represented.",
     "COMPLETE", "85%"),

    # ── PHASE 2 ─────────────────────────────────────────────────────────────
    ("P2.1", "Phase 2", "R1",
     "Gray reflection points in Figs 1E, 4C, 5C, 6B are confusing.",
     "Added caption sentence in all four figures: 'Gray points represent the symmetric counterpart of each coalescence event, included to illustrate symmetry.'",
     "COMPLETE", "95%"),

    ("P2.2", "Phase 2", "R1",
     "Fig. 2D: relationship between dots, squares, and gray bars is unclear.",
     "Expanded Fig. 2D caption to explicitly describe each element: dots=per-event correlations, squares=mean±s.e.m., gray lines=null model baseline, red/blue=same vs cross-community.",
     "COMPLETE", "95%"),

    ("P2.3", "Phase 2", "R1",
     "ED Fig. 5C: unclear whether mean markers are present.",
     "Investigated: means ARE present (red/blue squares); caption already describes them. No regeneration needed.",
     "COMPLETE", "90%"),

    ("P2.4", "Phase 2", "R2",
     "Improve pairwise correlation visualisation — group separation not visually clear.",
     "Documented 6 improvement options (larger markers, violin overlay, reduced dot opacity, Δ annotation). Figure regeneration recommended; final decision pending.",
     "PENDING DECISION", "80%"),

    # ── PHASE 3 ─────────────────────────────────────────────────────────────
    ("P3.1", "Phase 3", "R1",
     "Could absolute density (OD) differences explain Dominance outcomes?",
     "Tested 263 events: OD difference shows weak association (Spearman ρ=0.22, p<0.001) but denser parent wins only 27% of Dominance events (binomial p<0.0001). Rules out simple density explanation.",
     "COMPLETE", "90%"),

    ("P3.2", "Phase 3", "R1",
     "Does Dominance increase when parents differ in pH type (acidic vs alkaline)?",
     "Yes: different-pH pairs show 70.5% Dominance vs 54.1% for same-pH pairs (Fisher exact p=0.012). Supports pH-modification hypothesis as one contributing factor.",
     "COMPLETE", "95%"),

    ("P3.3", "Phase 3", "R1",
     "Fig. 5C PDI correlation may be circular if dominant species defines PDI.",
     "Removed dominant species and recomputed: R² drops 0.34→0.07, but 68% of events maintain the same winner direction (Spearman ρ=0.64, p<10⁻¹⁸), confirming subdominant community members also contribute.",
     "COMPLETE", "95%"),

    ("P3.4", "Phase 3", "R3",
     "Low richness geometrically inflates Dominance — need per-event additive null (n_C,null = n_A + n_B).",
     "Additive null classifies ALL 263 events as Mixing. 100% of observed Dominance events (157/157) deviate from null (Wilcoxon p=3.2×10⁻⁴⁴). Nutrient-driven gradient (39%→76%) is entirely absent from null. No geometric artifact.",
     "COMPLETE", "95%"),

    ("P3.5", "Phase 3", "R3",
     "Show species richness across Base/Nutr-/Nutr+ for synthetic communities.",
     "Richness does decrease with enrichment (KW p<0.001: Nutr- median=12, Base=8, Nutr+=7.5), but the Dominance shift (39%→76%) is disproportionately large, indicating a real ecological effect beyond geometry.",
     "COMPLETE", "95%"),

    ("P3.6", "Phase 3", "R2",
     "Classification depends on arbitrary thresholds; show continuous similarity measures.",
     "PDI (y) and retention magnitude (x²) distributions shown by medium. KW p=9×10⁻²⁵ for y; distribution shifts from bimodal (Nutr-) to strongly right-skewed (Nutr+), confirming categorical trends are threshold-independent.",
     "COMPLETE", "95%"),

    ("P3.7", "Phase 3", "R1",
     "Provide more detail on pool size effects on richness and Dominance.",
     "Realized richness increases with pool size (KW p=1×10⁻⁶); survival ratio decreases (larger pools filter more). Dominance frequency NOT significantly affected by pool size overall (χ²=2.24, p=0.69), but nutrient×pool-size interaction is evident.",
     "COMPLETE", "90%"),

    # ── PHASE 4 ─────────────────────────────────────────────────────────────
    ("P4.1", "Phase 4", "R3",
     "Increasing μ reduces richness, which geometrically inflates Dominance — show this confound is not the whole story.",
     "Composition-matched geometric null shows ~35% Dominance at high μ from unevenness alone, vs 77% observed (45% excess). Excess Dominance increases with μ (r=0.695, p<0.001), attributing the majority to ecological assembly structure.",
     "COMPLETE", "90%"),

    ("P4.2", "Phase 4", "R1",
     "Show interaction matrix block structure after assembly (within vs between community interactions).",
     "Within-community interactions (0.389±0.078) are consistently lower than between-community (0.500±0.046) at μ=0.50; ratio strengthens with μ. Assembly filters species into weakly-competing within-community groups.",
     "COMPLETE", "95%"),

    ("P4.3", "Phase 4", "R2",
     "Pairwise selection correlation is conceptually unclear; link to invasion fitness in gLV.",
     "Computed invasion concordance across μ values: excess concordance above independent null increases with μ (r=0.87, p<10⁻⁷). Co-assembled species share correlated invasion fates because assembly co-selects weakly-competing species.",
     "COMPLETE", "85%"),

    # ── PHASE 5 ─────────────────────────────────────────────────────────────
    ("P5.1", "Phase 5", "–",
     "Incorporate new analyses into results.tex.",
     "Not started — depends on all Phase 3/4 figures being finalised.",
     "NOT STARTED", "–"),

    ("P5.2", "Phase 5", "–",
     "Add new paragraphs to discussion.tex.",
     "Not started — depends on P1.7–P1.11 integration.",
     "NOT STARTED", "–"),

    ("P5.3", "Phase 5", "–",
     "Update methods.tex with pH details and gLV model framing.",
     "Not started.",
     "NOT STARTED", "–"),

    ("P5.4", "Phase 5", "–",
     "Update supplementary.tex with new figures and notes.",
     "Not started.",
     "NOT STARTED", "–"),

    ("P5.5", "Phase 5", "–",
     "Update figure captions for all new panels.",
     "Not started.",
     "NOT STARTED", "–"),

    ("P5.6", "Phase 5", "–",
     "Add new Extended Data / Supplementary figures.",
     "Not started — figures generated by Phase 3/4 scripts, need to be incorporated.",
     "NOT STARTED", "–"),

    ("P5.7", "Phase 5", "–",
     "Fix all cross-references after figure renumbering.",
     "Not started.",
     "NOT STARTED", "–"),

    # ── PHASE 6 ─────────────────────────────────────────────────────────────
    ("P6.1", "Phase 6", "R1",
     "Write point-by-point response to Reviewer 1.",
     "Not started — response_letter.tex does not exist yet.",
     "NOT STARTED", "–"),

    ("P6.2", "Phase 6", "R2",
     "Write point-by-point response to Reviewer 2.",
     "Not started.",
     "NOT STARTED", "–"),

    ("P6.3", "Phase 6", "R3",
     "Write point-by-point response to Reviewer 3.",
     "Not started.",
     "NOT STARTED", "–"),
]

HEADERS = ["ID", "Phase", "Reviewer", "Question summary", "Our answer / approach", "Status", "Confidence"]
COL_WIDTHS = [8, 12, 10, 52, 62, 18, 12]

# ── build workbook ───────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Revision Summary"

# header row
for col_i, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=1, column=col_i, value=header)
    cell.fill   = HEADER_FILL
    cell.font   = HEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col_i)].width = width

ws.row_dimensions[1].height = 22
ws.freeze_panes = "A2"

# data rows
for row_i, row_data in enumerate(ROWS, start=2):
    row_id, phase, reviewer, question, answer, status, confidence = row_data
    phase_hex = PHASE_COLORS.get(phase, "FFFFFF")
    status_hex = STATUS_COLORS.get(status, "FFFFFF")

    values = [row_id, phase, reviewer, question, answer, status, confidence]
    for col_i, val in enumerate(values, start=1):
        cell = ws.cell(row=row_i, column=col_i, value=val)
        cell.font   = BODY_FONT
        cell.border = BORDER
        # status column gets status colour, rest get phase colour
        if col_i == 6:
            cell.fill = PatternFill("solid", fgColor=status_hex)
            cell.alignment = CENTER_ALIGN
        elif col_i in (1, 2, 3, 7):
            cell.fill = PatternFill("solid", fgColor=phase_hex)
            cell.alignment = CENTER_ALIGN
        else:
            cell.fill = PatternFill("solid", fgColor=phase_hex)
            cell.alignment = WRAP_ALIGN
    ws.row_dimensions[row_i].height = 72

# auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# ── second sheet: legend ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Legend")
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 40

ws2.cell(1,1,"Phase colours").font = Font(bold=True)
for i, (phase, hex_) in enumerate(PHASE_COLORS.items(), start=2):
    ws2.cell(i, 1, phase).fill = PatternFill("solid", fgColor=hex_)
    ws2.cell(i, 1).border = BORDER
    ws2.cell(i, 2, f"Rows for {phase}").border = BORDER

ws2.cell(10,1,"Status colours").font = Font(bold=True)
for i, (st, hex_) in enumerate(STATUS_COLORS.items(), start=11):
    ws2.cell(i, 1, st).fill = PatternFill("solid", fgColor=hex_)
    ws2.cell(i, 1).border = BORDER
    ws2.cell(i, 2, st).border = BORDER

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Rows: {len(ROWS)} (+ 1 header)")
