"""
Generate revision_response_ordered.xlsx
Rows ordered by reviewer → point sequence as written in the review.
Columns: Review point | Reviewer | Section | Original question (verbatim excerpt) |
          Question summary | Our answer | Status | Confidence | Plan ID
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_PATH = (
    "/Users/jysong/Desktop/Gore_lab/Sequencing/Coalescence_session_20230404"
    "/Figure_generate/Draft/v5/latex/revision/revision_response_ordered.xlsx"
)

# ── styles ───────────────────────────────────────────────────────────────────
def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

HEADER_FILL  = fill("1A3A5C")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BODY_FONT    = Font(size=9)
WRAP         = Alignment(wrap_text=True, vertical="top")
CENTER       = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin         = Side(style="thin", color="CCCCCC")
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)

# per-reviewer row background
REV_FILL = {
    "R1":   "D6EAF8",   # blue
    "R2":   "D5F5E3",   # green
    "R3":   "FEF9E7",   # yellow
    "—":    "F2F3F4",   # grey (Phase 5/6)
}
STATUS_FILL = {
    "COMPLETE":         "A9DFBF",
    "IN PROGRESS":      "FAD7A0",
    "PENDING DECISION": "D7BDE2",
    "NOT STARTED":      "F5CBA7",
}
SECTION_FILL = {
    "Major":  "EBF5FB",
    "Minor":  "E9F7EF",
    "—":      "F9F9F9",
    "Manuscript integration": "F5EEF8",
    "Response letter":        "FDFEFE",
}

# ── data ─────────────────────────────────────────────────────────────────────
# Fields:
#   review_pt | reviewer | section
#   | original_quote (verbatim ~1 sentence from review)
#   | question_summary (our 1-2 sentence paraphrase)
#   | our_answer | status | confidence | plan_id

ROWS = [

    # ════════════════════════════════════════════════════════════════════
    # REVIEWER 1  (roughly descending order of importance)
    # ════════════════════════════════════════════════════════════════════
    ("R1-1", "R1", "Major",
     "…if community A reaches 10× the overall density of community B, then even "
     "relatively rare species in community A might be initially more abundant than "
     "the dominant species in community B… Could those [OD] data be used to check "
     "this possibility?",
     "Could absolute community density differences (OD) explain Dominance, rather than ecological competition?",
     "Tested 263 events: denser parent wins only 27% of Dominance events (binomial p<0.0001). "
     "OD difference shows only weak correlation with Dominance (Spearman ρ=0.22). Rules out simple density/dilution explanation.",
     "COMPLETE", "90%", "P3.1"),

    ("R1-2", "R1", "Major",
     "…it seems like an even more straightforward prediction would be that Dominance "
     "should become more likely when combining parent communities with qualitatively "
     "different pH (e.g. alkaline and acidic) compared to cases where both parents "
     "are either acidic or alkaline. Is this the case?",
     "Is Dominance more frequent when parents differ in pH type (acidic vs alkaline)?",
     "Yes: different-pH pairs → 70.5% Dominance vs 54.1% same-pH pairs (Fisher exact p=0.012). "
     "Continuous |ΔpH| analysis confirms a positive but modest association (Spearman ρ=0.15).",
     "COMPLETE", "95%", "P3.2"),

    ("R1-3", "R1", "Major",
     "In Fig. 5C, it seems like there is some risk of circularity if the dominant "
     "species are included in the calculation of the PDI. Does the positive correlation "
     "under Nutr+ conditions hold up if the dominant species are excluded?",
     "Does the Fig. 5C PDI correlation hold if the dominant species is removed (circularity check)?",
     "R² drops from 0.34 to 0.07 after removal, but 68% of events maintain the same winner direction "
     "(Spearman ρ=0.64, p<10⁻¹⁸) and 40% remain classified as Dominance, showing subdominant taxa also contribute.",
     "COMPLETE", "95%", "P3.3"),

    ("R1-4", "R1", "Major",
     "…could the Authors provide slightly more detail on the effect of the initial "
     "pool size… Is it similarly true that there is no significant effect of initial "
     "richness in the experimental communities? Perhaps by plotting the realized "
     "richness and interaction strength in parent communities as a function of "
     "initial richness? It might also be interesting to see the survival ratio.",
     "More detail on pool-size effects: realized richness, survival ratio, and Dominance frequency vs initial pool size.",
     "Richness increases with pool size (KW p=10⁻⁶) but Dominance frequency is NOT significantly affected (χ²=2.24, p=0.69). "
     "Survival ratio decreases with pool size. Nutrient×pool-size interaction is the more informative pattern.",
     "COMPLETE", "90%", "P3.7"),

    ("R1-5", "R1", "Major",
     "…'This pattern of Dominance as the most frequent outcome was robust across "
     "variants of similarity metrics'… Two of four alternative similarity metrics "
     "gave different results, with Dominance being the *least* likely outcome "
     "under Jensen-Shannon and Jaccard.",
     "Soften the claim that Dominance is 'robust' across similarity metrics.",
     "Revised to: 'broadly consistent across Bray–Curtis and Euclidean, but relative ranking of outcomes is "
     "sensitive to metric choice (Jensen–Shannon and Jaccard yield different orderings).'",
     "COMPLETE", "95%", "P1.5"),

    ("R1-6", "R1", "Minor",
     "It took me quite a while to realize that the gray points in Figs. 1E, 4C, "
     "5C, and 6B are simply a reflection of the colored points. Please indicate "
     "this clearly, or perhaps just remove these.",
     "Clarify in captions that gray points are symmetric reflections of each coalescence event.",
     "Added caption sentence to all four figures: 'Gray points represent the symmetric counterpart of each "
     "coalescence event (community B vs A), included to illustrate symmetry.'",
     "COMPLETE", "95%", "P2.1"),

    ("R1-7", "R1", "Minor",
     "In Fig. 2A, it might be helpful to show the matrix of interaction coefficients "
     "after assembly of the parent communities (with visible block structure).",
     "Show the interaction coefficient matrix post-assembly, revealing within- vs between-community block structure.",
     "Within-community interactions (mean=0.389) consistently lower than between-community (0.500) at μ=0.50. "
     "Block structure strengthens with μ. New supplementary figure generated.",
     "COMPLETE", "95%", "P4.2"),

    ("R1-8", "R1", "Minor",
     "Fig. 2D, I was somewhat confused by the relationship between point and squares "
     "(means?). Visually, the squares do not look like they are near the means of "
     "the points… the gray horizontal bars are not explained.",
     "Clarify Fig. 2D caption: explain dots, squares, gray bars, and red/blue color coding.",
     "Expanded caption: dots=per-event correlations, squares=mean±s.e.m., gray lines=shuffled-label null baseline, "
     "red=same-parent, blue=cross-community pairs.",
     "COMPLETE", "95%", "P2.2"),

    ("R1-9", "R1", "Minor",
     "I think it is worth emphasizing how interesting and perhaps surprising it is "
     "to find 'community-level cohesion without cooperation'… This finding will be "
     "counterintuitive to many readers.",
     "Expand the Tikhonov 'cohesion without cooperation' paragraph in the Discussion.",
     "Added 3-sentence expansion: purely competitive interactions, when structured by assembly history, "
     "are sufficient to produce correlated species fates and 'cohesion without cooperation'.",
     "COMPLETE", "90%", "P1.7"),

    ("R1-10", "R1", "Minor",
     "In ED Fig. 5C, are means missing?",
     "Confirm whether mean markers are present in ED Fig. 5C.",
     "Investigated: means ARE present (red/blue squares with error bars). Caption already describes them. "
     "No change needed.",
     "COMPLETE", "90%", "P2.3"),

    ("R1-11", "R1", "Minor",
     "On p.9 in the SI, I believe the sentence '…Extended Data Fig. 5' should refer "
     "to ED Fig. 4.",
     "Fix wrong Extended Data figure references in the Supplementary Information.",
     "Fixed two files: simulations.tex (ED Fig 5→4), pairwise_selection_correlation.tex (ED Fig 6→5, 7→6).",
     "COMPLETE", "100%", "P1.2"),

    # ════════════════════════════════════════════════════════════════════
    # REVIEWER 2  — Major comments
    # ════════════════════════════════════════════════════════════════════
    ("R2-1", "R2", "Major",
     "…the manuscript appears to conflate resource enrichment with increased "
     "interaction strength, whereas both theory and prior work suggest that "
     "enrichment restructures interaction networks in more complex ways… "
     "We therefore suggest reframing this result more broadly in terms of changes "
     "in interaction intensity and environmental feedbacks.",
     "Nutrient enrichment is not a simple proxy for increased pairwise interaction strength — reframe accordingly.",
     "Revised Results and Discussion to acknowledge nutrient enrichment as a complex perturbation (citing Duan 2025). "
     "Changed 'mean interaction strengths' → 'mean competition coefficients'; added note that the mapping is approximate.",
     "COMPLETE", "80%", "P1.11"),

    ("R2-2", "R2", "Major",
     "…the manuscript interprets dominance of one parental community and correlated "
     "persistence of its taxa as evidence for community-level selection… Similar "
     "patterns could arise from shared environmental filtering, correlated traits, "
     "or environmental modification… recommend discussing alternative mechanisms.",
     "Discuss alternative mechanisms (environmental filtering, hitchhiking, shared traits) that could mimic community-level selection.",
     "New Discussion paragraph addresses all three alternatives and argues against each using pairwise correlation controls, "
     "assembly history effects, and gLV reproduction of patterns without environmental filtering.",
     "COMPLETE", "85%", "P1.9"),

    ("R2-3", "R2", "Major",
     "…the classification into Dominance, Mixture, and Restructuring… depends on "
     "thresholding a continuous similarity measure… We suggest that the authors "
     "present continuous similarity measures alongside categorical outcomes.",
     "Present continuous PDI and retention magnitude distributions alongside the discrete categories.",
     "PDI (y) KDE distributions shown by medium: KW p=9×10⁻²⁵. Distribution shifts from bimodal (Nutr-) to "
     "strongly right-skewed (Nutr+), confirming trends are threshold-independent.",
     "COMPLETE", "95%", "P3.6"),

    ("R2-4", "R2", "Major",
     "…We suggest adding a short paragraph linking this metric more explicitly to "
     "invasion fitness in a gLV framework… Pairwise invasion assays can then be "
     "interpreted as empirical approximations of invasion fitness in two-species "
     "systems.",
     "Link pairwise selection correlation to invasion fitness in the gLV framework.",
     "Computed invasion concordance across μ: excess concordance above independent null increases with μ (r=0.87, p<10⁻⁷). "
     "Co-assembled species share correlated invasion fates because assembly co-selects weakly-competing taxa.",
     "COMPLETE", "85%", "P4.3"),

    ("R2-5", "R2", "Major",
     "…incubating diverse natural communities in simplified media is known to drive "
     "convergence toward a limited set of functional guilds… the 'natural' communities "
     "used here may already be filtered to resemble the synthetic communities… "
     "discuss the potential effects of this pre-selection more explicitly.",
     "Acknowledge potential pre-selection during the stabilisation phase for natural communities.",
     "New Discussion paragraph: 'stabilisation may pre-select species toward culture conditions, "
     "potentially reducing effective diversity' — acknowledged as a limitation that does not explain the qualitative trend.",
     "COMPLETE", "85%", "P1.10"),

    ("R2-6", "R2", "Major",
     "…We therefore suggest framing the model more explicitly as a phenomenological "
     "rather than mechanistic description of the system.",
     "Explicitly frame the gLV model as phenomenological, not mechanistic.",
     "Added sentence after model introduction: 'We emphasize that the gLV model serves as a phenomenological "
     "framework… rather than as a mechanistic model of specific biochemical interactions.'",
     "COMPLETE", "95%", "P1.6"),

    # ── Reviewer 2  — Minor comments ────────────────────────────────────────
    ("R2-M1", "R2", "Minor",
     "Improve the visualisation of pairwise correlation results, as the separation "
     "between groups is not visually clear.",
     "Improve Fig. 2D / ED Fig. 5 / ED Fig. 6 pairwise correlation plots for clearer group separation.",
     "Documented 6 improvement options (larger mean markers, violin overlay, reduced dot opacity, Δ annotation). "
     "Figure regeneration recommended; final approach not yet decided.",
     "PENDING DECISION", "80%", "P2.4"),

    ("R2-M2", "R2", "Minor",
     "Specify how pH was measured (e.g. continuously or at endpoints), and whether "
     "variability across replicates was assessed.",
     "Specify pH measurement method in Methods.",
     "Already specified in methods.tex (Apera Instruments PH5500 benchtop meter). No change needed.",
     "COMPLETE", "95%", "P1.12"),

    ("R2-M3", "R2", "Minor",
     "Clarify the biological interpretation of the interaction coefficient "
     "distributions used in the theoretical model.",
     "Clarify biological meaning of gLV interaction coefficients (α_ij) in the model description.",
     "Added to supplementary_methods.tex: α_ij encodes net per-capita effect including resource competition, "
     "interference, and metabolite-mediated inhibition; facilitation not represented (α_ij ≥ 0).",
     "COMPLETE", "85%", "P1.13"),

    ("R2-M4", "R2", "Minor",
     "Ensure consistent terminology and notation throughout (e.g. 'pairwise', "
     "'community-level', α_ij).",
     "Review and unify terminology and notation throughout the manuscript.",
     "Reviewed all .tex files — 'coalescence event/experiment', μ, and α_ij are used consistently. "
     "No changes required beyond the disambiguation sentence added for P1.4.",
     "COMPLETE", "95%", "P1.3"),

    ("R2-M5", "R2", "Minor",
     "Correct minor typographical issues, including 'generalis ability' → "
     "'generalisability'.",
     "Fix typo 'generalis ability' → 'generalisability'.",
     "Already corrected before the v5 copy (fixed during v3→v4 transition). No further change needed.",
     "COMPLETE", "100%", "P1.1"),

    # ════════════════════════════════════════════════════════════════════
    # REVIEWER 3
    # ════════════════════════════════════════════════════════════════════
    ("R3-1", "R3", "Major",
     "…it would be necessary to compare each coalescence community with a corresponding "
     "null model constructed specifically for each of the pairs of parental communities, "
     "such as the simple additive model n_C,null = n_A + n_B… on a case-by-case basis, "
     "not just by comparing distributions.",
     "Per-event additive null model: compare each observed n_C to n_C,null = normalize(n_A + n_B) individually.",
     "Additive null classifies ALL 263 events as Mixing. 100% of observed Dominance (157/157) deviates from null "
     "(Wilcoxon p=3.2×10⁻⁴⁴). Nutrient gradient (39%→76% Dominance) absent from null. No geometric artifact.",
     "COMPLETE", "95%", "P3.4"),

    ("R3-2a", "R3", "Major",
     "It would also be useful to see how richness (in the final communities, not in "
     "the inocula) changes with mu in the model, as well as across the different "
     "media (Base, Nutr-, Nutr+) used in the experiment.",
     "Show species richness across Base/Nutr-/Nutr+ for synthetic communities (experimental).",
     "Richness decreases with enrichment (KW p<0.001: Nutr- median=12, Base=8, Nutr+=7.5). "
     "The Dominance shift (39%→76%) is disproportionately large, indicating a real ecological effect beyond geometry.",
     "COMPLETE", "95%", "P3.5"),

    ("R3-2b", "R3", "Major",
     "It seems important to test whether the increase of 'Dominance' outcomes in "
     "the model as mu grows is substantially different from the expected increase in "
     "the null models as N decreases.",
     "Show richness vs μ in the model and test whether Dominance increase exceeds what richness reduction alone predicts.",
     "Composition-matched null accounts for ~35% Dominance at high μ from unevenness alone, vs 77% observed. "
     "Excess Dominance ≈45 percentage points increases with μ (r=0.695, p<0.001) — majority is ecological.",
     "COMPLETE", "90%", "P4.1"),

    ("R3-3", "R3", "Major",
     "…I would suggest the claims in section 2.6 of the Results as well as in the "
     "Discussion are toned down (e.g., lines 292-294) [regarding natural communities].",
     "Tone down claims about natural communities; acknowledge that facilitative interactions are not captured by the model.",
     "Softened language: 'higher Restructuring may reflect greater taxonomic diversity, more complex interaction networks, "
     "or facilitative interactions not captured by our competition-only model.' Caution against direct quantitative comparisons.",
     "COMPLETE", "90%", "P1.8"),

    ("R3-4", "R3", "Minor",
     "The term 'interaction strength' may be somewhat confusing since, in the context "
     "of the gLV model, it refers strictly to competitive interactions. Would two "
     "species which engage in a mutualism be considered 'weak interactors'? It may "
     "be worth considering using different terminology ('competition strength'?).",
     "Clarify 'interaction strength' refers to competitive interactions in the model context; consider alternate terminology.",
     "Added disambiguation sentence in results.tex at first use of μ: distinguishes model coefficient (competition) "
     "from experimental proxy (failed pairwise invasions). Term 'interaction strength' retained with qualifier.",
     "COMPLETE", "90%", "P1.4"),

    # ════════════════════════════════════════════════════════════════════
    # PHASE 5 — Manuscript integration (not reviewer-specific)
    # ════════════════════════════════════════════════════════════════════
    ("P5.1", "—", "Manuscript integration",
     "—",
     "Incorporate all new analyses and reframed language into results.tex.",
     "Not started — depends on all Phase 3/4 figures being finalised.",
     "NOT STARTED", "—", "P5.1"),

    ("P5.2", "—", "Manuscript integration",
     "—",
     "Add new Discussion paragraphs (alternative mechanisms, nutrient framing, pre-selection).",
     "Not started.",
     "NOT STARTED", "—", "P5.2"),

    ("P5.3", "—", "Manuscript integration",
     "—",
     "Update methods.tex with pH details and gLV model framing.",
     "Not started.",
     "NOT STARTED", "—", "P5.3"),

    ("P5.4", "—", "Manuscript integration",
     "—",
     "Update supplementary.tex with new figures and supplementary notes.",
     "Not started.",
     "NOT STARTED", "—", "P5.4"),

    ("P5.5", "—", "Manuscript integration",
     "—",
     "Update all figure captions for new panels.",
     "Not started.",
     "NOT STARTED", "—", "P5.5"),

    ("P5.6", "—", "Manuscript integration",
     "—",
     "Add new Extended Data / Supplementary figures from Phase 3/4 scripts.",
     "Not started — figures generated, need incorporation.",
     "NOT STARTED", "—", "P5.6"),

    ("P5.7", "—", "Manuscript integration",
     "—",
     "Fix all cross-references after figure renumbering.",
     "Not started.",
     "NOT STARTED", "—", "P5.7"),

    # ════════════════════════════════════════════════════════════════════
    # PHASE 6 — Response letter
    # ════════════════════════════════════════════════════════════════════
    ("P6.1", "R1", "Response letter",
     "—",
     "Write point-by-point response to Reviewer 1 (11 points).",
     "Not started — response_letter.tex does not exist yet.",
     "NOT STARTED", "—", "P6.1"),

    ("P6.2", "R2", "Response letter",
     "—",
     "Write point-by-point response to Reviewer 2 (6 major + 5 minor).",
     "Not started.",
     "NOT STARTED", "—", "P6.2"),

    ("P6.3", "R3", "Response letter",
     "—",
     "Write point-by-point response to Reviewer 3 (4 points).",
     "Not started.",
     "NOT STARTED", "—", "P6.3"),
]

HEADERS = [
    "Review\npoint", "Reviewer", "Section",
    "Original question (verbatim excerpt)",
    "Question summary (1–2 sentences)",
    "Our answer / approach",
    "Status", "Confidence", "Plan ID"
]
COL_WIDTHS = [9, 10, 22, 50, 42, 58, 20, 11, 8]

# ── build workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "By Review Order"

# header
for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = CENTER
    cell.border    = BORDER
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

# group separator rows
prev_reviewer = None
separator_rows = set()

for row_i, row_data in enumerate(ROWS, start=2):
    rev_pt, reviewer, section, orig_q, q_sum, answer, status, conf, plan_id = row_data
    rev_hex    = REV_FILL.get(reviewer, "FFFFFF")
    status_hex = STATUS_FILL.get(status, "FFFFFF")
    sec_hex    = SECTION_FILL.get(section, "FFFFFF")

    # insert a thin group-header row when reviewer changes
    if reviewer != prev_reviewer:
        # use current row_i, shift below data insertion handled by adjusting enumerate
        pass
    prev_reviewer = reviewer

    values = [rev_pt, reviewer, section, orig_q, q_sum, answer, status, conf, plan_id]
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row_i, column=ci, value=val)
        cell.font   = BODY_FONT
        cell.border = BORDER
        if ci == 7:                           # Status
            cell.fill      = PatternFill("solid", fgColor=status_hex)
            cell.alignment = CENTER
        elif ci in (1, 2, 8, 9):             # ID / Reviewer / Confidence / Plan ID
            cell.fill      = PatternFill("solid", fgColor=rev_hex)
            cell.alignment = CENTER
        elif ci == 3:                         # Section
            cell.fill      = PatternFill("solid", fgColor=sec_hex)
            cell.alignment = CENTER
        else:                                 # text columns
            cell.fill      = PatternFill("solid", fgColor=rev_hex)
            cell.alignment = WRAP
    ws.row_dimensions[row_i].height = 80

# auto-filter on header
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# ── Legend sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Legend")
ws2.column_dimensions["A"].width = 24
ws2.column_dimensions["B"].width = 38

legend_items = [
    ("Row background colours", None),
    ("Reviewer 1 rows",  "D6EAF8"),
    ("Reviewer 2 rows",  "D5F5E3"),
    ("Reviewer 3 rows",  "FEF9E7"),
    ("Phase 5/6 rows",   "F2F3F4"),
    ("", None),
    ("Status cell colours", None),
    ("COMPLETE",          "A9DFBF"),
    ("IN PROGRESS",       "FAD7A0"),
    ("PENDING DECISION",  "D7BDE2"),
    ("NOT STARTED",       "F5CBA7"),
    ("", None),
    ("Section colours", None),
    ("Major",                  "EBF5FB"),
    ("Minor",                  "E9F7EF"),
    ("Manuscript integration", "F5EEF8"),
    ("Response letter",        "FDFEFE"),
]
for ri, (label, hex_) in enumerate(legend_items, start=1):
    c = ws2.cell(ri, 1, label)
    if hex_:
        c.fill = PatternFill("solid", fgColor=hex_)
        c.border = BORDER
    else:
        c.font = Font(bold=True)

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Data rows: {len(ROWS)}")
