#!/usr/bin/env python3
"""
Build OD/pH distribution plots and an easy-to-use Excel workbook for Jiliang.

The data are spread across the processed community metadata, isolate plate
exports, and pairwise colony-count assays. Direct OD/pH exists for monoculture
isolates and community samples. I did not find a direct pairwise OD/pH workbook;
the available pairwise assay is CFU/count based, so this script exports that
separately and documents the gap.
"""

from __future__ import annotations

import re
from pathlib import Path

import numpy as np
import pandas as pd

import matplotlib as mpl
import matplotlib.pyplot as plt
import seaborn as sns

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = Path(__file__).resolve().parent
FIG_DIR = OUT_DIR / "figures"
OUT_XLSX = OUT_DIR / "To_Jiliang_OD_pH_distribution_20260605.xlsx"
OUT_MEMO = OUT_DIR / "README.md"

SAMPLE_SHEET = ROOT / "Postprocessed" / "Sample_Sheet.xlsx"
COMMUNITY_METADATA_LEGACY = ROOT / "Postprocessed" / "Metadata.xlsx"
ISOLATE_METADATA = ROOT / "Postprocessed" / "Metadata_Isolates.xlsx"
ISOLATE_PH = (
    ROOT
    / "ExperimentalResult"
    / "Data"
    / "2208_Coalescence_processed"
    / "pH_isolates"
    / "230623_pH.xlsx"
)
ISOLATE_OD = (
    ROOT
    / "ExperimentalResult"
    / "Data"
    / "2208_Coalescence_processed"
    / "pH_isolates"
    / "220910_54isolatesOD_flat100um.xlsx"
)
PAIRWISE_COUNTS = ROOT / "Postprocessed" / "PairwiseColonyCountings_processed_230915.xlsx"

MEDIUM_ORDER = ["Nutr-", "Base", "Nutr+"]
MEDIUM_FROM_PAIRWISE = {"LN": "Nutr-", "MN": "Base", "HN": "Nutr+"}
PAIRWISE_MEDIA = ["LN", "MN", "HN"]

sns.set_theme(style="ticks", context="paper")
mpl.rcParams["pdf.fonttype"] = 42
mpl.rcParams["ps.fonttype"] = 42
mpl.rcParams["font.family"] = "Arial"


def clean_numeric(value):
    """Convert plate-reader cells such as '48?' to numbers and blanks to NaN."""
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else np.nan


def find_plate_data_start(df: pd.DataFrame) -> int:
    for i in range(len(df)):
        if str(df.iloc[i, 0]).strip() == "A":
            return i
    raise ValueError("Could not find plate row A in Excel sheet.")


def load_plate_values(path: Path, sheet: str, value_name: str, scale: float = 1.0) -> pd.DataFrame:
    raw = pd.read_excel(path, sheet_name=sheet, header=None)
    start = find_plate_data_start(raw)
    rows = []
    for r, row_label in enumerate(list("ABCDEFGH")):
        if start + r >= len(raw):
            break
        for c in range(1, 13):
            value = clean_numeric(raw.iloc[start + r, c])
            if np.isfinite(value):
                rows.append(
                    {
                        "well": f"{row_label}{c}",
                        "plate_row": row_label,
                        "plate_col": c,
                        value_name: value / scale,
                    }
                )
    out = pd.DataFrame(rows)
    out["isolate_order"] = np.arange(1, len(out) + 1)
    return out


def load_single_species_isolates() -> pd.DataFrame:
    growth = pd.read_excel(ISOLATE_METADATA).head(54).copy()
    growth["isolate_order"] = np.arange(1, len(growth) + 1)
    growth["isolate_id"] = growth["isolate_order"].map(lambda x: f"isolate_{x:02d}")
    growth["well_growth_plate_order"] = [
        f"{row}{col}" for row in list("ABCDEF") for col in range(1, 10)
    ][: len(growth)]

    rename = {
        "LN_O1": "LN_growth_rate_expfit_h_inv",
        "LN_O2": "LN_growth_rate_15x_min_h_inv",
        "LN_O3": "LN_growth_rate_halfmax_h_inv",
        "LN_O4": "LN_max_growth_curve_OD",
        "MN_O1": "MN_growth_rate_expfit_h_inv",
        "MN_O2": "MN_growth_rate_15x_min_h_inv",
        "MN_O3": "MN_growth_rate_halfmax_h_inv",
        "MN_O4": "MN_max_growth_curve_OD",
        "HN_O1": "HN_growth_rate_expfit_h_inv",
        "HN_O2": "HN_growth_rate_15x_min_h_inv",
        "HN_O3": "HN_growth_rate_halfmax_h_inv",
        "HN_O4": "HN_max_growth_curve_OD",
    }
    growth = growth.rename(columns=rename)
    metric_cols = list(rename.values())
    for col in metric_cols:
        growth[col] = pd.to_numeric(growth[col], errors="coerce")
        growth.loc[growth[col] > 100, col] = np.nan

    ph = load_plate_values(ISOLATE_PH, "after 15h", "single_species_pH_after_15h", scale=10.0)
    od = load_plate_values(ISOLATE_OD, "Sheet4", "single_species_endpoint_OD", scale=1.0)
    ph = ph.head(54).rename(columns={"well": "well_pH_plate"})
    od = od.head(54).rename(columns={"well": "well_OD_plate"})

    out = growth.merge(
        ph[["isolate_order", "well_pH_plate", "single_species_pH_after_15h"]],
        on="isolate_order",
        how="left",
    )
    out = out.merge(
        od[["isolate_order", "well_OD_plate", "single_species_endpoint_OD"]],
        on="isolate_order",
        how="left",
    )
    out["pH_after_15h_minus_start_6p5"] = out["single_species_pH_after_15h"] - 6.5

    front = [
        "isolate_order",
        "isolate_id",
        "well_growth_plate_order",
        "well_pH_plate",
        "well_OD_plate",
        "single_species_pH_after_15h",
        "pH_after_15h_minus_start_6p5",
        "single_species_endpoint_OD",
    ]
    return out[front + metric_cols]


def load_community_samples() -> pd.DataFrame:
    raw = pd.read_excel(SAMPLE_SHEET, sheet_name="samples")
    df = raw.copy()
    df["sample_type_easy"] = df["sample_type"].replace(
        {"Subcommunity": "single_community", "Coalescence": "coalesced_community"}
    )

    for i in range(1, 8):
        df[f"OD_day{i}"] = pd.to_numeric(df[f"fieldOD{i}"], errors="coerce")
        df[f"pH_day{i}"] = pd.to_numeric(df[f"fieldPH{i}"], errors="coerce")

    df["OD_day7_minus_day1"] = df["OD_day7"] - df["OD_day1"]
    df["OD_mean_day5_to_day7"] = df[["OD_day5", "OD_day6", "OD_day7"]].mean(axis=1)
    df["OD_range_day5_to_day7"] = df[["OD_day5", "OD_day6", "OD_day7"]].max(axis=1) - df[
        ["OD_day5", "OD_day6", "OD_day7"]
    ].min(axis=1)
    df["max_endpoint_OD_day1_to_day7"] = df[[f"OD_day{i}" for i in range(1, 8)]].max(axis=1)
    df["pH_day7_minus_day1"] = df["pH_day7"] - df["pH_day1"]
    df["pH_mean_day5_to_day7"] = df[["pH_day5", "pH_day6", "pH_day7"]].mean(axis=1)
    df["pH_range_day5_to_day7"] = df[["pH_day5", "pH_day6", "pH_day7"]].max(axis=1) - df[
        ["pH_day5", "pH_day6", "pH_day7"]
    ].min(axis=1)

    ordered = [
        "sample_id",
        "community_origin",
        "medium",
        "sample_type_easy",
        "replicate",
        "community_idx",
        "parent_1_community_idx",
        "parent_2_community_idx",
        "parent_1_sample_id",
        "parent_2_sample_id",
        "OD_final_mean",
        "OD_final_std",
        "pH_final_mean",
        "pH_final_std",
        "growth_curve_AUC_mean",
        *[f"OD_day{i}" for i in range(1, 8)],
        "OD_day7_minus_day1",
        "OD_mean_day5_to_day7",
        "OD_range_day5_to_day7",
        "max_endpoint_OD_day1_to_day7",
        *[f"pH_day{i}" for i in range(1, 8)],
        "pH_day7_minus_day1",
        "pH_mean_day5_to_day7",
        "pH_range_day5_to_day7",
        "notes",
    ]
    return df[ordered].rename(columns={"sample_type_easy": "sample_type"})


def build_coalescence_parent_pairs(samples: pd.DataFrame) -> pd.DataFrame:
    coalesced = samples[samples["sample_type"] == "coalesced_community"].copy()
    lookup = samples.set_index("sample_id")
    rows = []
    for _, row in coalesced.iterrows():
        p1_id = row["parent_1_sample_id"]
        p2_id = row["parent_2_sample_id"]
        if pd.isna(p1_id) or pd.isna(p2_id) or p1_id not in lookup.index or p2_id not in lookup.index:
            continue
        p1 = lookup.loc[p1_id]
        p2 = lookup.loc[p2_id]
        rec = {
            "coalesced_sample_id": row["sample_id"],
            "community_origin": row["community_origin"],
            "medium": row["medium"],
            "replicate": row["replicate"],
            "coalesced_community_idx": row["community_idx"],
            "parent_1_sample_id": p1_id,
            "parent_2_sample_id": p2_id,
            "parent_1_community_idx": p1["community_idx"],
            "parent_2_community_idx": p2["community_idx"],
        }
        for metric in [
            "OD_day7",
            "OD_final_mean",
            "OD_mean_day5_to_day7",
            "max_endpoint_OD_day1_to_day7",
            "pH_day7",
            "pH_final_mean",
            "pH_mean_day5_to_day7",
        ]:
            rec[f"parent_1_{metric}"] = p1[metric]
            rec[f"parent_2_{metric}"] = p2[metric]
            rec[f"parent_pair_mean_{metric}"] = np.nanmean([p1[metric], p2[metric]])
            rec[f"parent_pair_abs_delta_{metric}"] = abs(p1[metric] - p2[metric])
            rec[f"coalesced_{metric}"] = row[metric]
            rec[f"coalesced_minus_parent_pair_mean_{metric}"] = row[metric] - rec[f"parent_pair_mean_{metric}"]
        rows.append(rec)
    return pd.DataFrame(rows)


def load_pairwise_medium(med: str):
    mono = pd.read_excel(PAIRWISE_COUNTS, sheet_name=f"{med}_mono")
    cc1 = pd.read_excel(PAIRWISE_COUNTS, sheet_name=f"{med}_1")
    cc2 = pd.read_excel(PAIRWISE_COUNTS, sheet_name=f"{med}_2")
    m_rep = mono.iloc[:, 1:].to_numpy(dtype=float)
    m = np.nanmean(m_rep, axis=0)
    cc1_arr = cc1.iloc[:, 1:].to_numpy(dtype=float)
    cc2_arr = cc2.iloc[:, 1:].to_numpy(dtype=float)
    return m_rep, m, cc1_arr, cc2_arr


def build_pairwise_coculture_counts(min_mono: float = 2.0):
    pair_rows = []
    species_rows = []
    for med in PAIRWISE_MEDIA:
        m_rep, m, cc1, cc2 = load_pairwise_medium(med)
        medium = MEDIUM_FROM_PAIRWISE[med]
        n = len(m)
        for i in range(n):
            for j in range(i + 1, n):
                mi, mj = m[i], m[j]
                if not np.isfinite(mi) or not np.isfinite(mj):
                    continue
                if mi < min_mono or mj < min_mono:
                    continue

                c_a_i, c_a_j = cc1[i, j], cc2[i, j]
                c_b_j, c_b_i = cc1[j, i], cc2[j, i]
                total_a = c_a_i + c_a_j if np.isfinite(c_a_i) and np.isfinite(c_a_j) else np.nan
                total_b = c_b_i + c_b_j if np.isfinite(c_b_i) and np.isfinite(c_b_j) else np.nan
                finite_totals = [x for x in [total_a, total_b] if np.isfinite(x)]
                if not finite_totals:
                    continue
                coculture_total = float(np.mean(finite_totals))
                additive = mi + mj
                ryt_a = (c_a_i / mi + c_a_j / mj) if np.isfinite(c_a_i) and np.isfinite(c_a_j) else np.nan
                ryt_b = (c_b_i / mi + c_b_j / mj) if np.isfinite(c_b_i) and np.isfinite(c_b_j) else np.nan
                finite_ryt = [x for x in [ryt_a, ryt_b] if np.isfinite(x)]
                ryt = float(np.mean(finite_ryt)) if finite_ryt else np.nan

                pair_rows.append(
                    {
                        "medium": medium,
                        "medium_code": med,
                        "species_i": i + 1,
                        "species_j": j + 1,
                        "mono_i_CFU_mean": mi,
                        "mono_j_CFU_mean": mj,
                        "mono_i_CFU_rep1": m_rep[0, i],
                        "mono_i_CFU_rep2": m_rep[1, i],
                        "mono_j_CFU_rep1": m_rep[0, j],
                        "mono_j_CFU_rep2": m_rep[1, j],
                        "additive_mono_sum_CFU": additive,
                        "coculture_total_i_resident_CFU": total_a,
                        "coculture_total_j_resident_CFU": total_b,
                        "coculture_total_mean_CFU": coculture_total,
                        "coculture_minus_additive_CFU": coculture_total - additive,
                        "coculture_minus_additive_fraction": (coculture_total - additive) / additive,
                        "relative_yield_total_RYT": ryt,
                        "subadditive_total": coculture_total < additive,
                        "RYT_below_1": ryt < 1 if np.isfinite(ryt) else np.nan,
                    }
                )

                for species, partner, mono_value, counts in [
                    (i + 1, j + 1, mi, [c_a_i, c_b_i]),
                    (j + 1, i + 1, mj, [c_a_j, c_b_j]),
                ]:
                    finite = [x for x in counts if np.isfinite(x)]
                    if not finite:
                        continue
                    coculture = float(np.mean(finite))
                    species_rows.append(
                        {
                            "medium": medium,
                            "medium_code": med,
                            "species": species,
                            "partner": partner,
                            "mono_CFU_mean": mono_value,
                            "coculture_CFU_mean_across_directions": coculture,
                            "relative_yield": coculture / mono_value if mono_value > 0 else np.nan,
                            "suppressed_in_coculture": coculture < mono_value,
                        }
                    )
    return pd.DataFrame(pair_rows), pd.DataFrame(species_rows)


def summarize_numeric(df: pd.DataFrame, group_cols: list[str], metrics: list[str], dataset: str) -> pd.DataFrame:
    rows = []
    for keys, sub in df.groupby(group_cols, dropna=False):
        if not isinstance(keys, tuple):
            keys = (keys,)
        key_data = dict(zip(group_cols, keys))
        for metric in metrics:
            values = pd.to_numeric(sub[metric], errors="coerce").dropna()
            rows.append(
                {
                    "dataset": dataset,
                    **key_data,
                    "metric": metric,
                    "n": len(values),
                    "mean": values.mean() if len(values) else np.nan,
                    "std": values.std() if len(values) > 1 else np.nan,
                    "median": values.median() if len(values) else np.nan,
                    "q25": values.quantile(0.25) if len(values) else np.nan,
                    "q75": values.quantile(0.75) if len(values) else np.nan,
                    "min": values.min() if len(values) else np.nan,
                    "max": values.max() if len(values) else np.nan,
                }
            )
    return pd.DataFrame(rows)


def build_summary(isolates, samples, parent_pairs, pairwise_counts, pairwise_species) -> pd.DataFrame:
    parts = [
        summarize_numeric(
            isolates.assign(group="all_single_species"),
            ["group"],
            ["single_species_endpoint_OD", "single_species_pH_after_15h", "pH_after_15h_minus_start_6p5"],
            "single_species",
        ),
        summarize_numeric(
            samples,
            ["community_origin", "medium", "sample_type"],
            ["OD_day7", "OD_final_mean", "OD_mean_day5_to_day7", "pH_day7", "pH_final_mean", "pH_mean_day5_to_day7"],
            "community_samples",
        ),
        summarize_numeric(
            parent_pairs,
            ["community_origin", "medium"],
            [
                "parent_pair_mean_OD_day7",
                "parent_pair_abs_delta_OD_day7",
                "coalesced_OD_day7",
                "parent_pair_mean_pH_day7",
                "parent_pair_abs_delta_pH_day7",
                "coalesced_pH_day7",
            ],
            "coalescence_parent_pairs",
        ),
        summarize_numeric(
            pairwise_counts,
            ["medium"],
            [
                "additive_mono_sum_CFU",
                "coculture_total_mean_CFU",
                "coculture_minus_additive_fraction",
                "relative_yield_total_RYT",
            ],
            "pairwise_coculture_CFU",
        ),
        summarize_numeric(
            pairwise_species,
            ["medium"],
            ["mono_CFU_mean", "coculture_CFU_mean_across_directions", "relative_yield"],
            "pairwise_species_CFU",
        ),
    ]
    return pd.concat(parts, ignore_index=True, sort=False)


def source_notes() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "item": "Community OD and pH, normalized sample sheet",
                "source": str(SAMPLE_SHEET.relative_to(ROOT)),
                "used_for": "community_OD_pH and coalescence_parent_pairs_OD_pH sheets",
                "notes": "One row per sequenced sample. Includes sample type, parent sample IDs for coalescence rows, OD_final_mean/std, pH_final_mean/std, and fieldOD/fieldPH day 1-7 values.",
            },
            {
                "item": "Community OD and pH, legacy processing output",
                "source": str(COMMUNITY_METADATA_LEGACY.relative_to(ROOT)),
                "used_for": "cross-check only",
                "notes": "Same underlying processed metadata used by earlier To_Jiliang export; Sample_Sheet.xlsx is cleaner because it includes parent IDs.",
            },
            {
                "item": "Single-species growth and max growth-curve OD",
                "source": str(ISOLATE_METADATA.relative_to(ROOT)),
                "used_for": "single_species_OD_pH sheet",
                "notes": "Generated by ExperimentalResult/IsolateExperimentalDataProcessing.m. O4 columns are max growth-curve OD by medium.",
            },
            {
                "item": "Single-species pH",
                "source": str(ISOLATE_PH.relative_to(ROOT)),
                "used_for": "single_species_OD_pH sheet",
                "notes": "Raw plate-reader pH values after 15 h; values are stored as pH x 10 and divided by 10 here.",
            },
            {
                "item": "Single-species endpoint OD",
                "source": str(ISOLATE_OD.relative_to(ROOT)),
                "used_for": "single_species_OD_pH sheet",
                "notes": "Raw plate-reader OD values from Sheet4, matching the earlier Jiliang export and existing monoculture histogram code.",
            },
            {
                "item": "Pairwise co-culture colony counts",
                "source": str(PAIRWISE_COUNTS.relative_to(ROOT)),
                "used_for": "pairwise_coculture_CFU and pairwise_species_CFU sheets",
                "notes": "This is CFU/count data, not OD/pH. It is the only direct pairwise assay table found in the workspace.",
            },
        ]
    )


def data_availability() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "level": "single_species",
                "OD_available": "yes",
                "pH_available": "yes",
                "source": str(ISOLATE_OD.relative_to(ROOT)) + "; " + str(ISOLATE_PH.relative_to(ROOT)),
                "comment": "Endpoint OD and pH after 15 h for 54 isolates.",
            },
            {
                "level": "pairwise_coculture",
                "OD_available": "not found",
                "pH_available": "not found",
                "source": str(PAIRWISE_COUNTS.relative_to(ROOT)),
                "comment": "Found pairwise CFU/count data for 12 isolates and 3 media; no direct pairwise OD/pH workbook was found by file and text search.",
            },
            {
                "level": "coalescence_parent_pair",
                "OD_available": "derived from parent community rows",
                "pH_available": "derived from parent community rows",
                "source": str(SAMPLE_SHEET.relative_to(ROOT)),
                "comment": "For each coalesced community, parent-1 and parent-2 OD/pH are joined from their single-community sample IDs.",
            },
            {
                "level": "community",
                "OD_available": "yes",
                "pH_available": "yes",
                "source": str(SAMPLE_SHEET.relative_to(ROOT)),
                "comment": "Day 1-7 endpoint OD and pH for single communities and coalesced communities.",
            },
        ]
    )


def save_fig(fig, basename: str) -> None:
    for ext in ["png", "pdf", "svg"]:
        fig.savefig(FIG_DIR / f"{basename}.{ext}", bbox_inches="tight", dpi=300)
    plt.close(fig)


def plot_single_species(isolates: pd.DataFrame) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(7.2, 3.0))
    sns.histplot(isolates["single_species_endpoint_OD"], kde=True, ax=axes[0], color="#2a9d8f")
    axes[0].set_xlabel("Single-species endpoint OD")
    axes[0].set_ylabel("Isolate count")
    sns.histplot(isolates["single_species_pH_after_15h"], bins=np.arange(2.75, 8.05, 0.25), ax=axes[1], color="#e76f51")
    axes[1].set_xlabel("Single-species pH after 15 h")
    axes[1].set_ylabel("Isolate count")
    fig.suptitle("Single-species OD and pH distributions", y=1.03)
    fig.tight_layout()
    save_fig(fig, "single_species_OD_pH_distributions")


def plot_community(samples: pd.DataFrame) -> None:
    plot_df = samples.copy()
    plot_df["medium"] = pd.Categorical(plot_df["medium"], categories=MEDIUM_ORDER, ordered=True)
    fig, axes = plt.subplots(2, 2, figsize=(8.2, 5.8), sharex=False)
    for ax, origin in zip(axes[:, 0], ["Synthetic", "Natural"]):
        sub = plot_df[plot_df["community_origin"] == origin]
        sns.boxplot(data=sub, x="medium", y="OD_day7", hue="sample_type", ax=ax, fliersize=1.8)
        ax.set_title(f"{origin}: OD day 7")
        ax.set_xlabel("")
        ax.set_ylabel("OD600")
        ax.legend(title="", fontsize=7)
    for ax, origin in zip(axes[:, 1], ["Synthetic", "Natural"]):
        sub = plot_df[plot_df["community_origin"] == origin]
        sns.boxplot(data=sub, x="medium", y="pH_day7", hue="sample_type", ax=ax, fliersize=1.8)
        ax.set_title(f"{origin}: pH day 7")
        ax.set_xlabel("")
        ax.set_ylabel("pH")
        ax.legend(title="", fontsize=7)
    fig.suptitle("Community-wise OD and pH distributions", y=1.02)
    fig.tight_layout()
    save_fig(fig, "community_OD_pH_distributions")


def plot_parent_pairs(parent_pairs: pd.DataFrame) -> None:
    plot_df = parent_pairs.copy()
    plot_df["medium"] = pd.Categorical(plot_df["medium"], categories=MEDIUM_ORDER, ordered=True)
    metrics = [
        ("parent_pair_mean_OD_day7", "Mean parent OD day 7"),
        ("parent_pair_abs_delta_OD_day7", "Abs parent OD difference"),
        ("coalesced_OD_day7", "Coalesced OD day 7"),
        ("parent_pair_mean_pH_day7", "Mean parent pH day 7"),
        ("parent_pair_abs_delta_pH_day7", "Abs parent pH difference"),
        ("coalesced_pH_day7", "Coalesced pH day 7"),
    ]
    fig, axes = plt.subplots(2, 3, figsize=(10.5, 5.8))
    for ax, (metric, label) in zip(axes.flat, metrics):
        sns.boxplot(data=plot_df, x="medium", y=metric, hue="community_origin", ax=ax, fliersize=1.8)
        ax.set_title(label)
        ax.set_xlabel("")
        ax.set_ylabel("")
        ax.legend(title="", fontsize=7)
    fig.suptitle("Coalescence parent-pair OD and pH context", y=1.02)
    fig.tight_layout()
    save_fig(fig, "coalescence_parent_pair_OD_pH_distributions")


def plot_pairwise_cfu(pairwise_counts: pd.DataFrame) -> None:
    plot_df = pairwise_counts.copy()
    plot_df["medium"] = pd.Categorical(plot_df["medium"], categories=MEDIUM_ORDER, ordered=True)
    long_totals = plot_df.melt(
        id_vars=["medium", "species_i", "species_j"],
        value_vars=["additive_mono_sum_CFU", "coculture_total_mean_CFU"],
        var_name="quantity",
        value_name="CFU",
    )
    long_totals["quantity"] = long_totals["quantity"].replace(
        {
            "additive_mono_sum_CFU": "monoculture sum",
            "coculture_total_mean_CFU": "coculture total",
        }
    )
    fig, axes = plt.subplots(1, 3, figsize=(10.5, 3.2))
    sns.boxplot(data=long_totals, x="medium", y="CFU", hue="quantity", ax=axes[0], fliersize=1.8)
    axes[0].set_title("Pairwise CFU totals")
    axes[0].set_xlabel("")
    axes[0].legend(title="", fontsize=7)
    sns.boxplot(data=plot_df, x="medium", y="coculture_minus_additive_fraction", ax=axes[1], color="#8ab17d", fliersize=1.8)
    axes[1].axhline(0, color="0.3", lw=0.8, ls="--")
    axes[1].set_title("Coculture minus additive")
    axes[1].set_xlabel("")
    axes[1].set_ylabel("fraction of additive")
    sns.boxplot(data=plot_df, x="medium", y="relative_yield_total_RYT", ax=axes[2], color="#f4a261", fliersize=1.8)
    axes[2].axhline(1, color="0.3", lw=0.8, ls="--")
    axes[2].set_title("Relative yield total")
    axes[2].set_xlabel("")
    axes[2].set_ylabel("RYT")
    fig.suptitle("Available pairwise assay distributions: CFU/counts, not OD/pH", y=1.04)
    fig.tight_layout()
    save_fig(fig, "pairwise_coculture_CFU_distributions")


def add_formatting(writer: pd.ExcelWriter) -> None:
    wb = writer.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"


def write_memo(isolates, samples, parent_pairs, pairwise_counts, pairwise_species) -> None:
    text = f"""# Jiliang OD/pH Distribution Export

Generated by:

```bash
python To_Jiliang_OD_pH_distribution_20260605/make_jiliang_od_ph_distribution.py
```

## Main outputs

- `To_Jiliang_OD_pH_distribution_20260605.xlsx`
- `figures/single_species_OD_pH_distributions.*`
- `figures/community_OD_pH_distributions.*`
- `figures/coalescence_parent_pair_OD_pH_distributions.*`
- `figures/pairwise_coculture_CFU_distributions.*`

## Data map

- Single-species OD: `{ISOLATE_OD.relative_to(ROOT)}`, sheet `Sheet4`.
- Single-species pH: `{ISOLATE_PH.relative_to(ROOT)}`, sheet `after 15h`, divided by 10.
- Community-wise OD/pH: `{SAMPLE_SHEET.relative_to(ROOT)}`, sheet `samples`, fields `fieldOD1`-`fieldOD7` and `fieldPH1`-`fieldPH7`.
- Coalescence parent-pair OD/pH: derived by joining each coalesced-community row to `parent_1_sample_id` and `parent_2_sample_id` in `{SAMPLE_SHEET.relative_to(ROOT)}`.
- Pairwise co-culture data found: `{PAIRWISE_COUNTS.relative_to(ROOT)}`. This is CFU/count data, not OD/pH.

## Important limitation

I searched file names and text references for pairwise/co-culture/invasion OD and pH sources. The only direct pairwise assay table found is the colony-count workbook above. Therefore the workbook includes pairwise CFU/count distributions and a separate availability sheet noting that direct pairwise OD/pH was not found in this workspace.

## Row counts

- Single-species isolates: {len(isolates)}
- Community samples: {len(samples)}
- Coalescence parent-pair rows: {len(parent_pairs)}
- Pairwise coculture pair rows: {len(pairwise_counts)}
- Pairwise species-in-pair rows: {len(pairwise_species)}
"""
    OUT_MEMO.write_text(text)


def main() -> None:
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    isolates = load_single_species_isolates()
    samples = load_community_samples()
    parent_pairs = build_coalescence_parent_pairs(samples)
    pairwise_counts, pairwise_species = build_pairwise_coculture_counts()
    summary = build_summary(isolates, samples, parent_pairs, pairwise_counts, pairwise_species)

    plot_single_species(isolates)
    plot_community(samples)
    plot_parent_pairs(parent_pairs)
    plot_pairwise_cfu(pairwise_counts)

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        source_notes().to_excel(writer, sheet_name="sources", index=False)
        data_availability().to_excel(writer, sheet_name="data_availability", index=False)
        isolates.to_excel(writer, sheet_name="single_species_OD_pH", index=False)
        samples.to_excel(writer, sheet_name="community_OD_pH", index=False)
        parent_pairs.to_excel(writer, sheet_name="parent_pairs_OD_pH", index=False)
        pairwise_counts.to_excel(writer, sheet_name="pairwise_coculture_CFU", index=False)
        pairwise_species.to_excel(writer, sheet_name="pairwise_species_CFU", index=False)
        summary.to_excel(writer, sheet_name="distribution_summary", index=False)
        add_formatting(writer)

    write_memo(isolates, samples, parent_pairs, pairwise_counts, pairwise_species)

    print(f"Wrote {OUT_XLSX}")
    print(f"Wrote {OUT_MEMO}")
    print(f"Wrote figures to {FIG_DIR}")
    print(f"single_species_OD_pH rows: {len(isolates)}")
    print(f"community_OD_pH rows: {len(samples)}")
    print(f"parent_pairs_OD_pH rows: {len(parent_pairs)}")
    print(f"pairwise_coculture_CFU rows: {len(pairwise_counts)}")


if __name__ == "__main__":
    main()
